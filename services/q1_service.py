#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Q1 职位差异度分析服务
"""

import logging
from functools import lru_cache
from pathlib import Path
from typing import List, Dict, Any, Optional

from openpyxl import load_workbook

from database.Q3 import DatabaseManager

logger = logging.getLogger(__name__)


class Q1Service:
    """Q1 职位差异度分析服务类"""
    
    def __init__(self, db_manager: DatabaseManager):
        self.db_manager = db_manager
    
    def get_representative_cities(self, limit: int = 30) -> List[str]:
        """
        获取30个代表性城市
        优先从 cluster_by_city 表中按职位数量排序选取前30个城市
        """
        # 使用 cluster_by_city 表（聚类后的城市-职位统计表）
        query = """
            SELECT city
            FROM cluster_by_city
            WHERE city IS NOT NULL
            GROUP BY city
            ORDER BY SUM(job_in_city_cnt) DESC
            LIMIT %s
        """
        
        results = self.db_manager.execute_query(query, (limit,))
        return [row[0] for row in results]
    
    def _parse_experience(self, experience: str) -> float:
        """
        将经验要求转换为数值
        例: "1-3年" -> 2, "3-5年" -> 4, "应届生" -> 0, "经验不限" -> 0
        """
        if not experience or experience in ['应届生', '经验不限', '在校生']:
            return 0
        
        # 处理"X年以上"的情况
        if '年以上' in experience:
            try:
                years = int(experience.replace('年以上', ''))
                return years + 2  # 例如："5年以上" -> 7
            except:
                return 0
        
        # 处理"X-Y年"的情况
        if '-' in experience and '年' in experience:
            try:
                parts = experience.replace('年', '').split('-')
                if len(parts) == 2:
                    start = int(parts[0])
                    end = int(parts[1])
                    return (start + end) / 2
            except:
                pass
        
        return 0
    
    def _parse_salary(self, salary: str) -> float:
        """
        计算薪资中位数/平均数
        例: "5-10K" -> 7.5, "10-15K" -> 12.5
        """
        if not salary:
            return 0
        
        try:
            # 移除"K"并按"-"分割
            salary = salary.upper().replace('K', '')
            if '-' in salary:
                parts = salary.split('-')
                if len(parts) == 2:
                    low = float(parts[0])
                    high = float(parts[1])
                    return (low + high) / 2
        except:
            pass
        
        return 0
    
    def get_scatter_data(self, city: str) -> Dict[str, Any]:
        """
        获取指定城市的散点气泡图数据（基于 cluster_by_city 表）
        
        - 以城市-职位聚合结果为基础（平均薪资 / 平均经验 / 平均学历等）
        - 使用 is_in_top200 标记该城市前200的职位
        - 追加来自原始 data 表的示例职位记录（最多3条）
        """
        # 1) 从 cluster_by_city 读取该城市的前200个职位（按 job_in_city_cnt 排序）
        cluster_query = """
            SELECT 
                job_title,
                avg_salary,
                salary_std,
                avg_education,
                avg_experience,
                min_annual_salary,
                max_annual_salary,
                job_in_city_cnt,
                is_in_top200,
                job_level_segment,
                avg_shannon_entropy
            FROM cluster_by_city
            WHERE city = %s
              AND is_in_top200 = 1
            ORDER BY job_in_city_cnt DESC
            LIMIT 200
        """
        cluster_rows = self.db_manager.execute_query(cluster_query, (city,))

        if not cluster_rows:
            return {
                "city": city,
                "total_jobs": 0,
                "data": []
            }

        # 收集职位名，用于后续到原始 data 表中查样本
        job_titles = [row[0] for row in cluster_rows]

        # 2) 从原始 data 表中获取该城市这些职位的示例记录（最多3条 / 职位）
        #    这样可以在前端 tooltip 中展示具体的薪资 / 学历 / 经验样本
        placeholders = ",".join(["%s"] * len(job_titles))
        samples_query = f"""
            SELECT 
                job_title,
                salary,
                education,
                experience,
                company_type
            FROM data
            WHERE city = %s
              AND job_title IN ({placeholders})
        """
        # 参数顺序：city + job_titles
        sample_params = (city, *job_titles)
        sample_rows = self.db_manager.execute_query(samples_query, sample_params)

        # 将样本记录按职位分组，并截取前3条
        samples_map: Dict[str, List[Dict[str, Any]]] = {}
        for job_title, salary, education, experience, company_type in sample_rows:
            if job_title not in samples_map:
                samples_map[job_title] = []
            if len(samples_map[job_title]) < 3:
                samples_map[job_title].append({
                    "salary": salary,
                    "education": education,
                    "experience": experience,
                    "company_type": company_type
                })

        # 3) 组装散点数据 & 归一化招聘人数（气泡大小）
        scatter_points: List[Dict[str, Any]] = []
        recruit_counts: List[int] = []

        for row in cluster_rows:
            (
                job_title,
                avg_salary,
                salary_std,
                avg_education,
                avg_experience,
                min_annual_salary,
                max_annual_salary,
                job_in_city_cnt,
                is_in_top200,
                job_level_segment,
                avg_shannon_entropy,
            ) = row

            # 确保job_in_city_cnt是整数，记录招聘人数用于归一化
            job_in_city_cnt = int(job_in_city_cnt) if job_in_city_cnt is not None else 0
            recruit_counts.append(job_in_city_cnt)

            # 选取该职位的一个代表 industry（行业）——从样本中取第一个的 company_type
            samples = samples_map.get(job_title, [])
            main_company_type = samples[0]["company_type"] if samples and samples[0]["company_type"] else "未知"

            scatter_points.append({
                "job_title": job_title,
                "city": city,
                # 聚合后的六个核心维度
                "avg_salary": float(avg_salary) if avg_salary is not None else 0.0,
                "salary_std": float(salary_std) if salary_std is not None else 0.0,
                "avg_education": float(avg_education) if avg_education is not None else 0.0,
                "avg_experience": float(avg_experience) if avg_experience is not None else 0.0,
                "min_annual_salary": float(min_annual_salary) if min_annual_salary is not None else 0.0,
                "max_annual_salary": float(max_annual_salary) if max_annual_salary is not None else 0.0,
                "job_in_city_cnt": job_in_city_cnt,
                "is_in_top200": bool(is_in_top200),
                "job_level": job_level_segment,          # 聚类后的职位层级
                "avg_shannon_entropy": float(avg_shannon_entropy) if avg_shannon_entropy is not None else 0.0,
                "company_type": main_company_type,
                # 示例列表，用于 tooltip 展示原始职位信息
                "samples": samples,
            })

        # 4) 计算归一化的招聘人数（用于气泡大小）
        if recruit_counts and len(recruit_counts) > 0:
            min_count = min(recruit_counts)
            max_count = max(recruit_counts)
            count_range = max_count - min_count if max_count > min_count else 1

            for point in scatter_points:
                count = int(point["job_in_city_cnt"]) if point["job_in_city_cnt"] is not None else 0
                # 归一化到0-1范围
                normalized = (count - min_count) / count_range if count_range > 0 else 0.5
                # 确保normalized在0-1范围内
                normalized = max(0.0, min(1.0, normalized))
                # 映射到合理的气泡大小范围（10-50）
                point["normalized_size"] = 10 + normalized * 40
        else:
            # 如果没有数据，给一个默认大小
            for point in scatter_points:
                point["normalized_size"] = 20

        return {
            "city": city,
            "total_jobs": len(scatter_points),
            "data": scatter_points
        }
    
    def get_job_levels(self) -> List[str]:
        """获取所有职位层级（聚类类别），基于 cluster_by_city 表的 job_level_segment 列"""
        query = """
            SELECT DISTINCT job_level_segment
            FROM cluster_by_city
            WHERE job_level_segment IS NOT NULL
            ORDER BY job_level_segment
        """
        results = self.db_manager.execute_query(query)
        return [row[0] for row in results]
    
    def get_industries(self, city: Optional[str] = None) -> List[str]:
        """获取行业类别"""
        if city:
            query = """
                SELECT DISTINCT company_type
                FROM data
                WHERE company_type IS NOT NULL
                    AND city = %s
                ORDER BY company_type
            """
            results = self.db_manager.execute_query(query, (city,))
        else:
            query = """
                SELECT DISTINCT company_type
                FROM data
                WHERE company_type IS NOT NULL
                ORDER BY company_type
            """
            results = self.db_manager.execute_query(query)
        
        return [row[0] for row in results]

    # --- 行业视图: 全国散点 ---

    _CITY_TIER_SCORE_MAP = {
        '一线': 4.0,
        '二线': 3.0,
        '三线': 2.0,
        '其他': 1.0
    }

    @staticmethod
    def _city_type_file() -> Path:
        return Path(__file__).resolve().parents[1] / 'dataset' / 'dataset' / '第四题' / 'city_type_statistics.xlsx'

    @staticmethod
    @lru_cache(maxsize=1)
    def _load_city_tier_cache() -> Dict[str, Dict[str, float]]:
        """读取 city_type_statistics.xlsx，缓存每个行业的加权城市等级"""
        cache: Dict[str, Dict[str, float]] = {}
        file_path = Q1Service._city_type_file()

        if not file_path.exists():
            logger.warning("city_type_statistics.xlsx 不存在，无法计算平均城市等级: %s", file_path)
            return cache

        try:
            workbook = load_workbook(file_path, read_only=True)
            sheet = workbook.active
            # 列顺序: city, city_tier, company_type, job_count, ...
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                city_tier = row[1]
                company_type = row[2]
                job_count = row[3]

                if not company_type or not city_tier or not job_count:
                    continue

                tier_score = Q1Service._CITY_TIER_SCORE_MAP.get(str(city_tier), 1.0)
                job_count = float(job_count)

                if company_type not in cache:
                    cache[company_type] = {
                        'weighted_score': 0.0,
                        'job_total': 0.0
                    }

                cache[company_type]['weighted_score'] += tier_score * job_count
                cache[company_type]['job_total'] += job_count
        except Exception as exc:
            logger.error("读取 city_type_statistics.xlsx 失败: %s", exc)
            return {}

        return cache

    @staticmethod
    def _score_to_tier_label(score: Optional[float]) -> Optional[str]:
        if score is None:
            return None
        if score >= 3.5:
            return '一线'
        if score >= 2.5:
            return '二线'
        if score >= 1.5:
            return '三线'
        return '其他'

    def get_national_industry_scatter(self) -> Dict[str, Any]:
        """
        获取全国行业散点数据
        - 直接使用 national_industry_stats 表
        - 追加 city_type_statistics.xlsx 的平均城市等级
        """
        query = """
            SELECT 
                company_type,
                national_job_count,
                avg_median_salary,
                avg_experience_rank,
                avg_education_rank
            FROM national_industry_stats
            WHERE company_type IS NOT NULL
        """

        rows = self.db_manager.execute_query(query)

        if not rows:
            return {
                "total_industries": 0,
                "data": []
            }

        tier_cache = self._load_city_tier_cache()
        scatter_points: List[Dict[str, Any]] = []
        job_counts: List[int] = []

        for row in rows:
            company_type = row[0]
            job_count = int(row[1]) if row[1] is not None else 0
            avg_salary = float(row[2]) if row[2] is not None else 0.0
            avg_experience = float(row[3]) if row[3] is not None else 0.0
            avg_education = float(row[4]) if row[4] is not None else 0.0

            tier_info = tier_cache.get(company_type)
            if tier_info and tier_info['job_total'] > 0:
                avg_city_tier_score = round(tier_info['weighted_score'] / tier_info['job_total'], 2)
            else:
                avg_city_tier_score = None

            scatter_points.append({
                "company_type": company_type,
                "job_count": job_count,
                "avg_median_salary": avg_salary,
                "avg_experience_rank": avg_experience,
                "avg_education_rank": avg_education,
                "avg_city_tier_score": avg_city_tier_score,
                "avg_city_tier_label": self._score_to_tier_label(avg_city_tier_score)
            })

            job_counts.append(job_count)

        # 归一化气泡大小（10-50）
        if job_counts:
            min_count = min(job_counts)
            max_count = max(job_counts)
            span = max_count - min_count if max_count > min_count else 1
            for item in scatter_points:
                normalized = (item["job_count"] - min_count) / span if span else 0.5
                normalized = max(0.0, min(1.0, normalized))
                item["normalized_size"] = 10 + normalized * 40
        else:
            for item in scatter_points:
                item["normalized_size"] = 20

        return {
            "total_industries": len(scatter_points),
            "data": scatter_points
        }

